"""Sincronização das notas CAPAG contra o Tesouro Transparente.

Não existe API — o Tesouro publica só arquivo estático (XLSX de municípios,
CSV de estados), atualizado ~3x/ano (ver docs/DOMINIO.md, seção CAPAG).
Cada sync busca a URL do recurso mais recente via `package_show` do CKAN
(nunca hardcoda nome de arquivo/data — eles mudam a cada publicação) e
grava por cima do que já existe (`bulk_create(update_conflicts=True)`,
mesmo padrão de `apps/catalogo/sync.py`).

Chamado por `apps/capag/tasks.py` (Celery Beat, mensal) e pelo management
command `sincronizar_capag` (manual, sem depender de Celery/RabbitMQ).
"""

from __future__ import annotations

import csv
import io
import logging

import httpx
from openpyxl import load_workbook

from .cores import NOTA_PARA_COR
from .models import EstadoCapag, MunicipioCapag

logger = logging.getLogger(__name__)

CKAN_PACKAGE_SHOW = "https://www.tesourotransparente.gov.br/ckan/api/3/action/package_show"
DATASET_MUNICIPIOS = "capag-municipios"
DATASET_ESTADOS = "capag-estados"
# Aba com a nota vigente (as outras abas do mesmo arquivo são detalhe por
# indicador/ano-base e não interessam aqui) — confirmado abrindo o arquivo
# real em 26/08/2026.
ABA_MUNICIPIOS = "Prévia da CAPAG"


def _url_do_recurso_mais_recente(client: httpx.Client, dataset: str) -> str:
    resp = client.get(CKAN_PACKAGE_SHOW, params={"id": dataset})
    resp.raise_for_status()
    recursos = resp.json()["result"]["resources"]
    return recursos[-1]["url"]


def _sincronizar_municipios(client: httpx.Client) -> int:
    url = _url_do_recurso_mais_recente(client, DATASET_MUNICIPIOS)
    conteudo = client.get(url).content

    planilha = load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    linhas = planilha[ABA_MUNICIPIOS].iter_rows(values_only=True)
    # As duas primeiras linhas são metadado da planilha (título mesclado e um
    # mapa de colunas auxiliar), não dado — o cabeçalho de verdade é a 3ª.
    next(linhas)
    next(linhas)
    indice = {nome: i for i, nome in enumerate(next(linhas))}

    registros = []
    for linha in linhas:
        codigo = linha[indice["Código Município Completo"]]
        nota = linha[indice["CAPAG"]]
        if not codigo or nota not in NOTA_PARA_COR:
            continue  # ente sem código ou não avaliado (`#N/A`/`n.d.`/`n.e.`)
        registros.append(
            MunicipioCapag(
                codigo_ibge=int(codigo),
                nome_municipio=linha[indice["Nome_Município"]] or "",
                uf=linha[indice["UF"]] or "",
                nota=nota,
            )
        )

    MunicipioCapag.objects.bulk_create(
        registros,
        update_conflicts=True,
        unique_fields=["codigo_ibge"],
        update_fields=["nome_municipio", "uf", "nota"],
    )
    return len(registros)


def _sincronizar_estados(client: httpx.Client) -> int:
    url = _url_do_recurso_mais_recente(client, DATASET_ESTADOS)
    conteudo = client.get(url).text

    leitor: csv.DictReader[str] = csv.DictReader(io.StringIO(conteudo), delimiter=";")
    registros = [
        EstadoCapag(uf=linha["UF"].strip().upper(), nota=linha["Classificação da CAPAG"].strip())
        for linha in leitor
        if linha.get("UF") and linha.get("Classificação da CAPAG", "").strip() in NOTA_PARA_COR
    ]

    EstadoCapag.objects.bulk_create(
        registros,
        update_conflicts=True,
        unique_fields=["uf"],
        update_fields=["nota"],
    )
    return len(registros)


def sincronizar(client: httpx.Client | None = None) -> dict[str, int]:
    """Baixa e grava as duas tabelas. Retorna quantos registros gravou em cada."""

    proprio = client is None
    http = client or httpx.Client(timeout=60.0, follow_redirects=True)
    try:
        municipios = _sincronizar_municipios(http)
        estados = _sincronizar_estados(http)
    finally:
        if proprio:
            http.close()

    logger.info("CAPAG sincronizado: %s municípios, %s estados", municipios, estados)
    return {"municipios": municipios, "estados": estados}
