"""Testes dos endpoints do Cotador (`/api/cotador/`).

O que este arquivo mais protege é a regra de produto que o Cotador
introduz: **salvar a cotação é o que salva a oportunidade** — quem abre o
modal a partir da busca, mexe e desiste não deixa nada na lista da equipe.

Precisam de banco (`manage.py test apps.cotador`). A conta em si está em
`test_formulas.py`, que roda sem Django.
"""

from __future__ import annotations

from decimal import Decimal

from rest_framework.test import APITestCase

from apps.accounts.models import User
from apps.fornecedores.models import Fornecedor
from apps.licitacoes.models import EventoOportunidadeSalva, OportunidadeSalva
from apps.licitacoes.test_salvas import item as item_da_busca

from .models import Cotacao


def oferta(**overrides) -> dict:
    base = {"nome": "Distribuidora Sul", "custo_produto": "24.90", "frete": "1.20", "outros": "0"}
    return {**base, **overrides}


def item_cotado(**overrides) -> dict:
    base = {
        "numero_item": "1",
        "descricao": "Papel A4 75g — resma 500fl",
        "unidade": "RESMA",
        "quantidade": "120",
        "valor_referencia": "30.00",
        "margem_minima": None,
        "margem_maxima": None,
        "impostos": None,
        "ofertas": [oferta(escolhida=True)],
    }
    return {**base, **overrides}


def cotacao_payload(**overrides) -> dict:
    base = {
        "titulo": "Cotação Prefeitura de Campinas",
        "transporte": "8",
        "garantia": "0",
        "lucro_minimo": "10",
        "lucro_maximo": "35",
        "impostos": "10",
        "itens": [item_cotado()],
    }
    return {**base, **overrides}


class SalvarCotacaoDaBuscaTests(APITestCase):
    """O caminho "Cotar" do card de oportunidade pesquisada."""

    def setUp(self):
        self.user = User.objects.create_user(email="operador@empresa.com", password="x")
        self.client.force_authenticate(self.user)

    def test_salvar_a_cotacao_salva_a_oportunidade_junto(self):
        resposta = self.client.post(
            "/api/cotador/cotacoes/",
            cotacao_payload(oportunidade={"itens": [item_da_busca()]}),
            format="json",
        )

        self.assertEqual(resposta.status_code, 201)
        self.assertTrue(resposta.data["oportunidade_criada"])
        self.assertEqual(OportunidadeSalva.objects.ativas().count(), 1)
        self.assertEqual(Cotacao.objects.count(), 1)

    def test_nao_salvar_a_cotacao_nao_deixa_nada_na_lista(self):
        """Não há rascunho persistido: sem POST, nada existe. O teste
        documenta a regra — não há endpoint que crie a oportunidade ao
        *abrir* o Cotador."""

        self.assertEqual(OportunidadeSalva.objects.count(), 0)
        self.assertEqual(Cotacao.objects.count(), 0)

    def test_salvar_de_novo_sobrescreve_a_mesma_cotacao(self):
        primeira = self.client.post(
            "/api/cotador/cotacoes/",
            cotacao_payload(oportunidade={"itens": [item_da_busca()]}),
            format="json",
        )

        segunda = self.client.post(
            "/api/cotador/cotacoes/",
            cotacao_payload(
                oportunidade={"itens": [item_da_busca()]},
                itens=[item_cotado(descricao="Papel A4 — outra marca")],
            ),
            format="json",
        )

        self.assertEqual(segunda.status_code, 200)
        self.assertEqual(segunda.data["id"], primeira.data["id"])
        self.assertEqual(Cotacao.objects.count(), 1)
        self.assertEqual(OportunidadeSalva.objects.ativas().count(), 1)
        self.assertFalse(segunda.data["oportunidade_criada"])

    def test_registra_a_cotacao_no_historico_da_oportunidade(self):
        self.client.post(
            "/api/cotador/cotacoes/",
            cotacao_payload(oportunidade={"itens": [item_da_busca()]}),
            format="json",
        )

        salva = OportunidadeSalva.objects.get()
        tipos = list(salva.eventos.values_list("tipo", flat=True))
        self.assertIn(EventoOportunidadeSalva.Tipo.SALVA, tipos)
        self.assertIn(EventoOportunidadeSalva.Tipo.PROPOSTA_GERADA, tipos)

    def test_sem_oportunidade_nenhuma_e_400(self):
        resposta = self.client.post("/api/cotador/cotacoes/", cotacao_payload(), format="json")
        self.assertEqual(resposta.status_code, 400)

    def test_exige_autenticacao(self):
        self.client.force_authenticate(None)
        self.assertEqual(
            self.client.post("/api/cotador/cotacoes/", cotacao_payload(), format="json").status_code,
            401,
        )


class CotacaoDeUmaSalvaTests(APITestCase):
    """O caminho "Abrir cotação" do módulo Oportunidades / Salvas."""

    def setUp(self):
        self.user = User.objects.create_user(email="operador@empresa.com", password="x")
        self.client.force_authenticate(self.user)
        self.salva = OportunidadeSalva.objects.create(
            cnpj_orgao="12345678000199",
            ano_compra="2026",
            sequencial_compra="42",
            objeto="Aquisição de café e açúcar",
            itens=[item_da_busca()],
        )

    def salvar(self, **overrides):
        return self.client.post(
            "/api/cotador/cotacoes/",
            cotacao_payload(oportunidade_id=self.salva.pk, **overrides),
            format="json",
        )

    def test_oportunidade_ja_salva_nao_vira_duplicata(self):
        resposta = self.salvar()

        self.assertEqual(resposta.status_code, 201)
        self.assertFalse(resposta.data["oportunidade_criada"])
        self.assertEqual(OportunidadeSalva.objects.count(), 1)

    def test_sem_cotacao_o_endpoint_da_oportunidade_devolve_404(self):
        """É o sinal de que o modal abre em branco, com os itens do
        snapshot — não é erro a mostrar pro usuário."""

        resposta = self.client.get(f"/api/cotador/oportunidades/{self.salva.pk}/cotacao/")
        self.assertEqual(resposta.status_code, 404)

    def test_abre_a_cotacao_pela_oportunidade(self):
        self.salvar()

        resposta = self.client.get(f"/api/cotador/oportunidades/{self.salva.pk}/cotacao/")

        self.assertEqual(resposta.status_code, 200)
        self.assertEqual(resposta.data["titulo"], "Cotação Prefeitura de Campinas")
        self.assertEqual(len(resposta.data["itens"]), 1)

    def test_excluir_a_cotacao_mantem_a_oportunidade_salva(self):
        cotacao_id = self.salvar().data["id"]

        self.assertEqual(
            self.client.delete(f"/api/cotador/cotacoes/{cotacao_id}/").status_code, 204
        )
        self.assertFalse(Cotacao.objects.exists())
        self.assertTrue(OportunidadeSalva.objects.ativas().exists())


class TotaisEValidacaoTests(APITestCase):
    def setUp(self):
        self.user = User.objects.create_user(email="operador@empresa.com", password="x")
        self.client.force_authenticate(self.user)
        self.salva = OportunidadeSalva.objects.create(
            cnpj_orgao="12345678000199", ano_compra="2026", sequencial_compra="42"
        )

    def salvar(self, **overrides):
        return self.client.post(
            "/api/cotador/cotacoes/",
            cotacao_payload(oportunidade_id=self.salva.pk, **overrides),
            format="json",
        )

    def test_total_do_cliente_e_ignorado_e_recalculado_no_servidor(self):
        resposta = self.salvar(valor_cotado="1.00", lucro_total="0.01")

        # 26,10 × (1+8%+35%) ÷ 90% × 120 unidades = 4.976,40
        self.assertEqual(Decimal(resposta.data["valor_cotado"]), Decimal("4976.40"))
        self.assertEqual(Decimal(resposta.data["lucro_total"]), Decimal("1096.20"))

    def test_resposta_traz_totais_e_derivados_por_item(self):
        dados = self.salvar().data

        self.assertEqual(dados["totais"]["itens"], 1)
        self.assertEqual(dados["totais"]["pendencias"], 0)
        self.assertEqual(len(dados["itens_calculados"]), 1)
        self.assertGreater(
            Decimal(dados["itens_calculados"][0]["preco_final_unitario"]),
            Decimal(dados["itens_calculados"][0]["preco_reserva_unitario"]),
        )

    def test_item_sem_fornecedor_nenhum_e_recusado(self):
        resposta = self.salvar(itens=[item_cotado(ofertas=[])])
        self.assertEqual(resposta.status_code, 400)

    def test_dois_fornecedores_escolhidos_no_mesmo_item_sao_recusados(self):
        resposta = self.salvar(
            itens=[item_cotado(ofertas=[oferta(escolhida=True), oferta(escolhida=True)])]
        )
        self.assertEqual(resposta.status_code, 400)

    def test_sem_ninguem_marcado_a_primeira_oferta_e_gravada_como_escolhida(self):
        """O banco conta a mesma história que a tela (que já usa a primeira
        na conta)."""

        self.salvar(itens=[item_cotado(ofertas=[oferta(), oferta(nome="Outro")])])

        ofertas = Cotacao.objects.get().itens.get().ofertas.all()
        self.assertEqual([o.escolhida for o in ofertas], [True, False])

    def test_lucro_minimo_acima_do_maximo_e_recusado(self):
        resposta = self.salvar(lucro_minimo="50", lucro_maximo="10")
        self.assertEqual(resposta.status_code, 400)

    def test_percentual_absurdo_e_recusado(self):
        self.assertEqual(self.salvar(transporte="900").status_code, 400)

    def test_fornecedor_do_cadastro_e_vinculado_e_o_nome_vem_dele(self):
        fornecedor = Fornecedor.objects.create(
            nome="Distribuidora Sul Ltda", cnpj="11222333000181", email="v@sul.com"
        )

        self.salvar(
            itens=[
                item_cotado(
                    ofertas=[oferta(fornecedor=fornecedor.pk, nome="nome antigo", escolhida=True)]
                )
            ]
        )

        gravada = Cotacao.objects.get().itens.get().ofertas.get()
        self.assertEqual(gravada.fornecedor, fornecedor)
        self.assertEqual(gravada.nome, "Distribuidora Sul Ltda")

    def test_excluir_o_fornecedor_nao_apaga_de_quem_a_equipe_cotou(self):
        fornecedor = Fornecedor.objects.create(
            nome="Distribuidora Sul Ltda", cnpj="11222333000181", email="v@sul.com"
        )
        self.salvar(
            itens=[item_cotado(ofertas=[oferta(fornecedor=fornecedor.pk, escolhida=True)])]
        )

        fornecedor.delete()

        gravada = Cotacao.objects.get().itens.get().ofertas.get()
        self.assertIsNone(gravada.fornecedor)
        self.assertEqual(gravada.nome, "Distribuidora Sul Ltda")


class PlanilhaTests(APITestCase):
    def setUp(self):
        self.user = User.objects.create_user(email="operador@empresa.com", password="x")
        self.client.force_authenticate(self.user)
        salva = OportunidadeSalva.objects.create(
            cnpj_orgao="12345678000199",
            ano_compra="2026",
            sequencial_compra="42",
            objeto="Aquisição de papel",
            uasg="925997",
        )
        self.cotacao_id = self.client.post(
            "/api/cotador/cotacoes/",
            cotacao_payload(oportunidade_id=salva.pk),
            format="json",
        ).data["id"]

    def test_exporta_xlsx_com_nome_de_arquivo(self):
        resposta = self.client.get(f"/api/cotador/cotacoes/{self.cotacao_id}/planilha/")

        self.assertEqual(resposta.status_code, 200)
        self.assertIn("spreadsheetml", resposta["Content-Type"])
        self.assertIn("925997", resposta["Content-Disposition"])

    def test_planilha_sai_com_formulas_e_nao_com_numeros_congelados(self):
        from openpyxl import load_workbook
        from io import BytesIO

        resposta = self.client.get(f"/api/cotador/cotacoes/{self.cotacao_id}/planilha/")
        workbook = load_workbook(BytesIO(resposta.content))

        self.assertEqual(
            [aba.title for aba in workbook.worksheets], ["Proposta", "Comparativo", "Parâmetros"]
        )
        # Os nomes definidos são o que faz mudar um parâmetro recalcular a
        # proposta inteira.
        for nome in ("Transporte", "Garantia", "LucroMinimo", "LucroMaximo", "Tributos"):
            self.assertIn(nome, workbook.defined_names)

        proposta = workbook["Proposta"]
        formulas = [
            celula.value
            for linha in proposta.iter_rows()
            for celula in linha
            if isinstance(celula.value, str) and celula.value.startswith("=")
        ]
        self.assertTrue(any("Transporte" in f for f in formulas))
        self.assertTrue(any(f.startswith("=SUM(") for f in formulas))
