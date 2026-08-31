"""A planilha de proposta que o botão "Exportar" do Cotador gera.

**Não é um dump da tela — é uma planilha que continua funcionando.** Toda
coluna calculada sai como *fórmula*, não como número congelado: quem receber
o arquivo pode mexer no custo de um fornecedor, na margem ou na carga
tributária e ver preço, lucro e totais se refazerem sozinhos. Os percentuais
da cotação viram *nomes definidos* (`Transporte`, `LucroMaximo`, ...) na aba
"Parâmetros", então trocar um único valor lá recalcula a proposta inteira.

Isso também é o que torna a planilha conferível: a conta que o sistema fez
está visível na barra de fórmulas, não escondida no servidor.

**Item com regra própria vira literal; item sem regra vira referência.** Um
item que usa a margem padrão da cotação recebe a fórmula `=LucroMaximo`, e
acompanha a mudança do parâmetro. Um item com margem própria recebe o
número — porque foi uma decisão específica daquele item e não pode ser
apagada por uma mudança global.

Três abas:

- **Proposta** — o documento em si: cabeçalho com a marca, dados do edital,
  a tabela de itens e o resumo financeiro.
- **Comparativo** — todos os fornecedores cotados por item, com o escolhido
  e o mais barato marcados. É a justificativa da escolha.
- **Parâmetros** — os percentuais, com a explicação da divisão por
  `(1 − tributos)`.
"""

from __future__ import annotations

import datetime as dt
from decimal import Decimal
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ImagemPlanilha
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.worksheet import Worksheet

from apps.fornecedores.documentos import formatar_documento

from .formulas import calcular_item, melhor, totalizar
from .models import Cotacao

LOGO = Path(__file__).parent / "assets" / "logo.png"

# Mesmos tokens de `frontend/src/styles/_tokens.scss` — a planilha é o
# documento que sai da empresa, e tem que parecer com o produto.
NAVY = "FF16294D"
AZUL = "FF2B7CC4"
AZUL_SUAVE = "FFEAF2FA"
BORDA = "FFDBE3EC"
CINZA_CLARO = "FFF7F9FC"
MUTED = "FF66788F"
VERDE = "FF2F7D4F"
AMBAR = "FFB8791D"
VERMELHO = "FFC0392B"

MOEDA = 'R$ #,##0.00'
MOEDA_4 = 'R$ #,##0.0000'
QUANTIDADE = '#,##0.####'
# O valor gravado é 35 (pontos percentuais), não 0,35 — as fórmulas dividem
# por 100, igual à tela. O formato só acrescenta o símbolo.
PERCENTUAL = '0.00"%"'
DESCONTO = '0.0%'

FINA = Side(style="thin", color=BORDA)
GRADE = Border(left=FINA, right=FINA, top=FINA, bottom=FINA)


def _titulo(celula, *, tamanho=11, cor="FF16294D", negrito=True):
    celula.font = Font(name="Calibri", size=tamanho, bold=negrito, color=cor)
    return celula


def _cabecalho_tabela(aba: Worksheet, linha: int, titulos: list[str]) -> None:
    for coluna, texto in enumerate(titulos, start=1):
        celula = aba.cell(row=linha, column=coluna, value=texto)
        celula.font = Font(name="Calibri", size=9, bold=True, color="FFFFFFFF")
        celula.fill = PatternFill("solid", fgColor=NAVY)
        celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celula.border = GRADE
    aba.row_dimensions[linha].height = 30


def _larguras(aba: Worksheet, larguras: dict[str, int]) -> None:
    for coluna, largura in larguras.items():
        aba.column_dimensions[coluna].width = largura


def _rotulo_valor(aba: Worksheet, linha: int, rotulo: str, valor: object) -> int:
    """Uma linha "Rótulo: valor" dos blocos de identificação. Devolve a
    próxima linha, para o chamador não ficar contando."""

    celula = aba.cell(row=linha, column=1, value=rotulo)
    celula.font = Font(name="Calibri", size=9, bold=True, color=MUTED)
    celula.alignment = Alignment(vertical="top")

    destino = aba.cell(row=linha, column=2, value=valor)
    destino.font = Font(name="Calibri", size=10, color="FF16294D")
    destino.alignment = Alignment(vertical="top", wrap_text=True)
    return linha + 1


def gerar_planilha(cotacao: Cotacao) -> bytes:
    """Devolve o .xlsx da proposta como bytes, pronto para virar resposta
    HTTP."""

    workbook = Workbook()
    parametros = _aba_parametros(workbook, cotacao)
    proposta = workbook.active
    proposta.title = "Proposta"
    _aba_proposta(proposta, cotacao)
    _aba_comparativo(workbook, cotacao)

    # Parâmetros vai por último na ordem das abas: é referência, não o que
    # se abre primeiro.
    workbook.move_sheet(parametros, offset=len(workbook.worksheets))
    workbook.active = 0

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def nome_do_arquivo(cotacao: Cotacao) -> str:
    oportunidade = cotacao.oportunidade
    partes = ["proposta", oportunidade.uasg or oportunidade.cnpj_orgao, oportunidade.ano_compra]
    base = "-".join(p for p in partes if p)
    return f"{base}-{dt.date.today():%Y%m%d}.xlsx"


# --------------------------------------------------------------------------
# Aba "Parâmetros" — a fonte dos nomes definidos que a Proposta referencia.
# --------------------------------------------------------------------------

# (rótulo, atributo do model, nome definido, ajuda)
_PARAMETROS = [
    ("Transporte", "transporte", "Transporte", "% sobre o custo"),
    ("Garantia extra", "garantia", "Garantia", "% sobre o custo"),
    ("Lucro mínimo", "lucro_minimo", "LucroMinimo", "% sobre o custo — define o preço de reserva"),
    ("Lucro máximo (alvo)", "lucro_maximo", "LucroMaximo", "% sobre o custo — define o preço proposto"),
    ("Tributos", "impostos", "Tributos", "% sobre a venda — ICMS/Simples, PIS, COFINS, IPI e ISS somados"),
]


def _aba_parametros(workbook: Workbook, cotacao: Cotacao) -> Worksheet:
    aba = workbook.create_sheet("Parâmetros")
    _larguras(aba, {"A": 26, "B": 14, "C": 62})

    _titulo(aba.cell(row=1, column=1, value="Parâmetros da cotação"), tamanho=13)
    aba.cell(
        row=2,
        column=1,
        value="Mude um valor aqui e a aba Proposta se recalcula — exceto nos itens que têm regra própria.",
    ).font = Font(name="Calibri", size=9, italic=True, color=MUTED)

    _cabecalho_tabela(aba, 4, ["Parâmetro", "Valor", "Base de cálculo"])

    for indice, (rotulo, atributo, nome, ajuda) in enumerate(_PARAMETROS):
        linha = 5 + indice
        aba.cell(row=linha, column=1, value=rotulo).font = Font(name="Calibri", size=10, bold=True)
        valor = aba.cell(row=linha, column=2, value=getattr(cotacao, atributo))
        valor.number_format = PERCENTUAL
        valor.alignment = Alignment(horizontal="center")
        valor.fill = PatternFill("solid", fgColor=AZUL_SUAVE)
        aba.cell(row=linha, column=3, value=ajuda).font = Font(name="Calibri", size=9, color=MUTED)
        for coluna in range(1, 4):
            aba.cell(row=linha, column=coluna).border = GRADE

        workbook.defined_names[nome] = DefinedName(
            nome, attr_text=f"'{aba.title}'!${get_column_letter(2)}${linha}"
        )

    explicacao = 5 + len(_PARAMETROS) + 1
    _titulo(aba.cell(row=explicacao, column=1, value="Por que o preço é uma divisão"), tamanho=11)
    aba.cell(
        row=explicacao + 1,
        column=1,
        value=(
            "Transporte, garantia e lucro incidem sobre o CUSTO; os tributos incidem sobre a "
            "VENDA — que é justamente o que se quer descobrir. Dividir o custo majorado por "
            "(1 − tributos) embute o imposto que vai cair sobre o próprio preço."
        ),
    ).alignment = Alignment(wrap_text=True, vertical="top")
    aba.merge_cells(start_row=explicacao + 1, start_column=1, end_row=explicacao + 3, end_column=3)

    formula = explicacao + 5
    aba.cell(
        row=formula,
        column=1,
        value="Preço unitário = Custo un. × (1 + (Transporte + Garantia)/100 + Margem/100) ÷ (1 − Tributos/100)",
    ).font = Font(name="Consolas", size=9, color="FF16294D")

    aba.sheet_view.showGridLines = False
    return aba


# --------------------------------------------------------------------------
# Aba "Proposta"
# --------------------------------------------------------------------------

COLUNAS_PROPOSTA = [
    "Item",
    "Descrição",
    "Unid.",
    "Qtd.",
    "Fornecedor",
    "Custo produto (un.)",
    "Frete (un.)",
    "Outros (un.)",
    "Custo un.",
    "Custo total",
    "Margem mín. %",
    "Margem máx. %",
    "Tributos %",
    "Preço un.",
    "Preço de reserva un.",
    "Lucro un.",
    "Lucro total",
    "Total do item",
    "Ref. edital (un.)",
    "Desconto s/ ref.",
]

# Índice (1-based) de cada coluna, pelo nome — evita constante mágica nas
# fórmulas e mantém tudo alinhado se uma coluna nova entrar no meio.
COL = {nome: indice for indice, nome in enumerate(COLUNAS_PROPOSTA, start=1)}


def _c(nome: str, linha: int) -> str:
    return f"{get_column_letter(COL[nome])}{linha}"


def _aba_proposta(aba: Worksheet, cotacao: Cotacao) -> None:
    oportunidade = cotacao.oportunidade
    itens = list(cotacao.itens.prefetch_related("ofertas"))

    _larguras(
        aba,
        {
            "A": 6,
            "B": 46,
            "C": 8,
            "D": 10,
            "E": 24,
            "F": 15,
            "G": 12,
            "H": 12,
            "I": 13,
            "J": 14,
            "K": 13,
            "L": 13,
            "M": 11,
            "N": 14,
            "O": 17,
            "P": 13,
            "Q": 14,
            "R": 15,
            "S": 15,
            "T": 14,
        },
    )

    linha = _cabecalho_documento(aba, cotacao)
    linha = _bloco_edital(aba, oportunidade, linha)

    primeira = linha + 1
    _cabecalho_tabela(aba, primeira, COLUNAS_PROPOSTA)

    padroes = cotacao.padroes
    for posicao, item in enumerate(itens):
        _linha_item(aba, primeira + 1 + posicao, item, padroes)

    ultima = primeira + len(itens)
    totais_linha = _linha_totais(aba, primeira, ultima, quantos=len(itens))
    _bloco_resumo(aba, cotacao, itens, totais_linha + 2)

    # Cabeçalho da tabela sempre visível ao rolar os itens.
    aba.freeze_panes = aba.cell(row=primeira + 1, column=1)
    if itens:
        aba.auto_filter.ref = f"A{primeira}:{get_column_letter(len(COLUNAS_PROPOSTA))}{ultima}"

    aba.sheet_view.showGridLines = False
    _configurar_impressao(aba)


def _cabecalho_documento(aba: Worksheet, cotacao: Cotacao) -> int:
    """Faixa da marca + título do documento. Devolve a linha seguinte."""

    for coluna in range(1, len(COLUNAS_PROPOSTA) + 1):
        for linha in range(1, 6):
            aba.cell(row=linha, column=coluna).fill = PatternFill("solid", fgColor=NAVY)
    for linha, altura in ((1, 12), (2, 20), (3, 20), (4, 18), (5, 10)):
        aba.row_dimensions[linha].height = altura

    if LOGO.exists():
        logo = ImagemPlanilha(str(LOGO))
        # ~52px de altura: cabe nas cinco linhas da faixa sem esticá-la.
        logo.height = 52
        logo.width = 52
        aba.add_image(logo, "A2")

    marca = aba.cell(row=2, column=2, value="INSIDE solutions")
    marca.font = Font(name="Calibri", size=15, bold=True, color="FFFFFFFF")
    marca.alignment = Alignment(vertical="center")

    sub = aba.cell(row=3, column=2, value="Proposta comercial · formação de preço")
    sub.font = Font(name="Calibri", size=10, color="FFBFD4EC")
    sub.alignment = Alignment(vertical="center")

    titulo = cotacao.titulo or (cotacao.oportunidade.objeto or "Cotação")
    identificacao = aba.cell(row=2, column=COL["Preço un."], value=titulo[:120])
    identificacao.font = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
    identificacao.alignment = Alignment(horizontal="right", vertical="center")

    emissao = aba.cell(
        row=3,
        column=COL["Preço un."],
        value=f"Emitida em {dt.date.today():%d/%m/%Y}",
    )
    emissao.font = Font(name="Calibri", size=9, color="FFBFD4EC")
    emissao.alignment = Alignment(horizontal="right", vertical="center")

    for linha in (2, 3):
        aba.merge_cells(
            start_row=linha,
            start_column=COL["Preço un."],
            end_row=linha,
            end_column=len(COLUNAS_PROPOSTA),
        )

    return 7


def _bloco_edital(aba: Worksheet, oportunidade, linha: int) -> int:
    _titulo(aba.cell(row=linha, column=1, value="Dados do edital"), tamanho=12)
    linha += 1

    prazo = oportunidade.data_encerramento_proposta
    dados = [
        ("Órgão", oportunidade.orgao_nome or "—"),
        ("UASG", oportunidade.uasg or "—"),
        ("Município / UF", f"{oportunidade.municipio or '—'} / {oportunidade.uf or '—'}"),
        ("Modalidade", oportunidade.modalidade or "—"),
        ("Propostas até", f"{prazo:%d/%m/%Y}" if prazo else "não informado"),
        ("Objeto", (oportunidade.objeto or "—").strip()),
    ]
    if oportunidade.link_pncp:
        dados.append(("Publicação (PNCP)", oportunidade.link_pncp))

    for rotulo, valor in dados:
        aba.merge_cells(start_row=linha, start_column=2, end_row=linha, end_column=8)
        linha = _rotulo_valor(aba, linha, rotulo, valor)

    return linha


def _linha_item(aba: Worksheet, linha: int, item, padroes) -> None:
    calculo = calcular_item(_para_formula(item), padroes)
    oferta = calculo.escolhida

    aba.cell(row=linha, column=COL["Item"], value=item.numero_item or str(item.ordem + 1))
    aba.cell(row=linha, column=COL["Descrição"], value=item.descricao or "—")
    aba.cell(row=linha, column=COL["Unid."], value=item.unidade or "—")
    aba.cell(row=linha, column=COL["Qtd."], value=item.quantidade)
    aba.cell(row=linha, column=COL["Fornecedor"], value=(oferta.nome if oferta else "") or "—")

    aba.cell(
        row=linha,
        column=COL["Custo produto (un.)"],
        value=oferta.custo_produto if oferta else Decimal("0"),
    )
    aba.cell(row=linha, column=COL["Frete (un.)"], value=oferta.frete if oferta else Decimal("0"))
    aba.cell(row=linha, column=COL["Outros (un.)"], value=oferta.outros if oferta else Decimal("0"))

    # A partir daqui é tudo fórmula — ver a docstring do módulo.
    aba.cell(
        row=linha,
        column=COL["Custo un."],
        value=f"={_c('Custo produto (un.)', linha)}+{_c('Frete (un.)', linha)}+{_c('Outros (un.)', linha)}",
    )
    aba.cell(
        row=linha,
        column=COL["Custo total"],
        value=f"={_c('Custo un.', linha)}*{_c('Qtd.', linha)}",
    )

    # Item sem regra própria referencia o parâmetro (e acompanha a mudança);
    # item com regra própria vira número (decisão específica dele).
    aba.cell(
        row=linha,
        column=COL["Margem mín. %"],
        value="=LucroMinimo" if item.margem_minima is None else item.margem_minima,
    )
    aba.cell(
        row=linha,
        column=COL["Margem máx. %"],
        value="=LucroMaximo" if item.margem_maxima is None else item.margem_maxima,
    )
    aba.cell(
        row=linha,
        column=COL["Tributos %"],
        value="=Tributos" if item.impostos is None else item.impostos,
    )

    majoracao = "(1+(Transporte+Garantia)/100+{margem}/100)"
    divisor = f"(1-{_c('Tributos %', linha)}/100)"
    aba.cell(
        row=linha,
        column=COL["Preço un."],
        value=(
            f"={_c('Custo un.', linha)}*"
            f"{majoracao.format(margem=_c('Margem máx. %', linha))}/{divisor}"
        ),
    )
    aba.cell(
        row=linha,
        column=COL["Preço de reserva un."],
        value=(
            f"={_c('Custo un.', linha)}*"
            f"{majoracao.format(margem=_c('Margem mín. %', linha))}/{divisor}"
        ),
    )
    aba.cell(
        row=linha,
        column=COL["Lucro un."],
        value=f"={_c('Custo un.', linha)}*{_c('Margem máx. %', linha)}/100",
    )
    aba.cell(
        row=linha,
        column=COL["Lucro total"],
        value=f"={_c('Lucro un.', linha)}*{_c('Qtd.', linha)}",
    )
    aba.cell(
        row=linha,
        column=COL["Total do item"],
        value=f"={_c('Preço un.', linha)}*{_c('Qtd.', linha)}",
    )

    aba.cell(row=linha, column=COL["Ref. edital (un.)"], value=item.valor_referencia)
    # Sem referência publicada, a célula fica vazia em vez de mostrar 100%
    # de desconto sobre zero.
    aba.cell(
        row=linha,
        column=COL["Desconto s/ ref."],
        value=(
            f'=IF({_c("Ref. edital (un.)", linha)}>0,'
            f'1-{_c("Preço un.", linha)}/{_c("Ref. edital (un.)", linha)},"")'
        ),
    )

    _formatar_linha_item(aba, linha, destacar=calculo.incompleto)


def _formatar_linha_item(aba: Worksheet, linha: int, *, destacar: bool) -> None:
    zebra = PatternFill("solid", fgColor=CINZA_CLARO) if linha % 2 == 0 else None
    alerta = PatternFill("solid", fgColor="FFFDF0EE")

    for coluna in range(1, len(COLUNAS_PROPOSTA) + 1):
        celula = aba.cell(row=linha, column=coluna)
        celula.border = GRADE
        celula.font = Font(name="Calibri", size=10)
        if destacar:
            celula.fill = alerta
        elif zebra:
            celula.fill = zebra

    aba.cell(row=linha, column=COL["Descrição"]).alignment = Alignment(
        wrap_text=True, vertical="top"
    )
    aba.cell(row=linha, column=COL["Item"]).alignment = Alignment(horizontal="center")
    aba.cell(row=linha, column=COL["Unid."]).alignment = Alignment(horizontal="center")
    aba.cell(row=linha, column=COL["Qtd."]).number_format = QUANTIDADE

    for nome in ("Custo produto (un.)", "Frete (un.)", "Outros (un.)"):
        aba.cell(row=linha, column=COL[nome]).number_format = MOEDA_4
    for nome in (
        "Custo un.",
        "Custo total",
        "Preço de reserva un.",
        "Lucro un.",
        "Lucro total",
        "Total do item",
        "Ref. edital (un.)",
    ):
        aba.cell(row=linha, column=COL[nome]).number_format = MOEDA
    for nome in ("Margem mín. %", "Margem máx. %", "Tributos %"):
        celula = aba.cell(row=linha, column=COL[nome])
        celula.number_format = PERCENTUAL
        celula.alignment = Alignment(horizontal="center")

    preco = aba.cell(row=linha, column=COL["Preço un."])
    preco.number_format = MOEDA
    preco.font = Font(name="Calibri", size=10, bold=True, color="FF16294D")

    aba.cell(row=linha, column=COL["Preço de reserva un."]).font = Font(
        name="Calibri", size=10, color=AMBAR
    )
    for nome in ("Lucro un.", "Lucro total"):
        aba.cell(row=linha, column=COL[nome]).font = Font(name="Calibri", size=10, color=VERDE)

    aba.cell(row=linha, column=COL["Desconto s/ ref."]).number_format = DESCONTO
    aba.row_dimensions[linha].height = 28


def _linha_totais(aba: Worksheet, cabecalho: int, ultima: int, *, quantos: int) -> int:
    linha = ultima + 1
    primeira = cabecalho + 1

    rotulo = aba.cell(row=linha, column=1, value="TOTAL")
    rotulo.font = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
    aba.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=COL["Fornecedor"])

    somas = {
        "Custo total": VERMELHO,
        "Lucro total": VERDE,
        "Total do item": "FFFFFFFF",
    }
    for nome, cor in somas.items():
        coluna = get_column_letter(COL[nome])
        celula = aba.cell(
            row=linha,
            column=COL[nome],
            # Cotação sem item nenhum não pode virar `SUM(A9:A8)`.
            value=f"=SUM({coluna}{primeira}:{coluna}{ultima})" if quantos else 0,
        )
        celula.number_format = MOEDA
        celula.font = Font(name="Calibri", size=11, bold=True, color=cor)

    for coluna in range(1, len(COLUNAS_PROPOSTA) + 1):
        celula = aba.cell(row=linha, column=coluna)
        celula.fill = PatternFill("solid", fgColor=NAVY)
        celula.border = GRADE
        if not celula.font.bold:
            celula.font = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
    aba.row_dimensions[linha].height = 24

    return linha


def _bloco_resumo(aba: Worksheet, cotacao: Cotacao, itens: list, linha: int) -> None:
    """O fecho da proposta: o que a equipe olha antes de dar o lance.

    Valores calculados no servidor (não fórmula): são leitura de resultado,
    e repetir aqui as somas da tabela só criaria um segundo lugar para
    divergir.
    """

    totais = totalizar([_para_formula(item) for item in itens], cotacao.padroes)

    _titulo(aba.cell(row=linha, column=1, value="Resumo financeiro"), tamanho=12)
    linha += 1

    blocos = [
        ("Custo dos produtos", totais.custo_produtos, MOEDA, "FF16294D", "sem frete e extras"),
        ("Capital necessário", totais.capital, MOEDA, "FF16294D", "com frete e extras"),
        ("Transporte", totais.transporte, MOEDA, "FF16294D", f"{cotacao.transporte}% do custo"),
        ("Impostos embutidos", totais.impostos, MOEDA, "FF16294D", "sobre a venda"),
        ("Valor cotado", totais.valor_cotado, MOEDA, NAVY, "preço proposto"),
        ("Preço de reserva", totais.preco_reserva, MOEDA, AMBAR, "piso: menor preço que compensa"),
        ("Folga para negociar", totais.folga, MOEDA, "FF16294D", "do preço proposto até o piso"),
        ("Lucro", totais.lucro_total, MOEDA, VERDE, f"{totais.lucro_percentual:.1f}% da venda"),
        (
            "Margem média",
            totais.margem_media / 100,
            "0.0%",
            VERDE,
            "lucro sobre o capital",
        ),
    ]

    for rotulo, valor, formato, cor, ajuda in blocos:
        celula_rotulo = aba.cell(row=linha, column=1, value=rotulo)
        celula_rotulo.font = Font(name="Calibri", size=10, bold=True, color=MUTED)

        celula_valor = aba.cell(row=linha, column=2, value=valor)
        celula_valor.number_format = formato
        celula_valor.font = Font(name="Calibri", size=11, bold=True, color=cor)

        aba.cell(row=linha, column=3, value=ajuda).font = Font(
            name="Calibri", size=9, italic=True, color=MUTED
        )
        aba.merge_cells(start_row=linha, start_column=3, end_row=linha, end_column=6)
        linha += 1

    if totais.pendencias:
        aviso = aba.cell(
            row=linha + 1,
            column=1,
            value=(
                f"Atenção: {totais.pendencias} item(ns) sem descrição ou sem preço de "
                "fornecedor — as linhas destacadas em vermelho."
            ),
        )
        aviso.font = Font(name="Calibri", size=10, bold=True, color=VERMELHO)


# --------------------------------------------------------------------------
# Aba "Comparativo"
# --------------------------------------------------------------------------

COLUNAS_COMPARATIVO = [
    "Item",
    "Descrição",
    "Fornecedor",
    "CNPJ / CPF",
    "Situação",
    "Custo produto (un.)",
    "Frete (un.)",
    "Outros (un.)",
    "Custo un.",
    "Escolhido",
    "Diferença p/ o mais barato",
]


def _aba_comparativo(workbook: Workbook, cotacao: Cotacao) -> Worksheet:
    aba = workbook.create_sheet("Comparativo")
    _larguras(
        aba,
        {"A": 7, "B": 40, "C": 26, "D": 20, "E": 20, "F": 17, "G": 12, "H": 12, "I": 14, "J": 11, "K": 24},
    )

    _titulo(aba.cell(row=1, column=1, value="Comparativo de fornecedores"), tamanho=13)
    aba.cell(
        row=2,
        column=1,
        value="Todos os preços recebidos por item. O escolhido é o que entra na proposta; o mais barato aparece em verde.",
    ).font = Font(name="Calibri", size=9, italic=True, color=MUTED)

    _cabecalho_tabela(aba, 4, COLUNAS_COMPARATIVO)
    linha = 5

    for item in cotacao.itens.prefetch_related("ofertas__fornecedor"):
        estrutura = _para_formula(item)
        mais_barata = melhor(estrutura)
        melhor_custo = mais_barata.custo_unitario if mais_barata else Decimal("0")

        for oferta in item.ofertas.all():
            fornecedor = oferta.fornecedor
            e_melhor = bool(mais_barata and mais_barata.identificador == oferta.pk)

            aba.cell(row=linha, column=1, value=item.numero_item or str(item.ordem + 1))
            aba.cell(row=linha, column=2, value=item.descricao or "—")
            aba.cell(row=linha, column=3, value=oferta.nome or "—")
            aba.cell(
                row=linha,
                column=4,
                value=formatar_documento(fornecedor.cnpj) if fornecedor else "—",
            )
            aba.cell(
                row=linha,
                column=5,
                # Fornecedor apagado do cadastro depois da cotação: o nome
                # sobrevive no snapshot, a situação não existe mais.
                value=fornecedor.get_situacao_display() if fornecedor else "fora do cadastro",
            )
            aba.cell(row=linha, column=6, value=oferta.custo_produto)
            aba.cell(row=linha, column=7, value=oferta.frete)
            aba.cell(row=linha, column=8, value=oferta.outros)
            aba.cell(row=linha, column=9, value=f"=F{linha}+G{linha}+H{linha}")
            aba.cell(row=linha, column=10, value="✓" if oferta.escolhida else "")
            aba.cell(
                row=linha,
                column=11,
                value=(
                    Decimal("0")
                    if e_melhor or not oferta.custo_produto
                    else oferta.custo_unitario - melhor_custo
                ),
            )

            _formatar_linha_comparativo(aba, linha, escolhida=oferta.escolhida, melhor=e_melhor)
            linha += 1

    aba.freeze_panes = aba.cell(row=5, column=1)
    if linha > 5:
        aba.auto_filter.ref = f"A4:{get_column_letter(len(COLUNAS_COMPARATIVO))}{linha - 1}"
    aba.sheet_view.showGridLines = False
    _configurar_impressao(aba)
    return aba


def _formatar_linha_comparativo(
    aba: Worksheet, linha: int, *, escolhida: bool, melhor: bool
) -> None:
    fundo = PatternFill("solid", fgColor=AZUL_SUAVE) if escolhida else None

    for coluna in range(1, len(COLUNAS_COMPARATIVO) + 1):
        celula = aba.cell(row=linha, column=coluna)
        celula.border = GRADE
        celula.font = Font(name="Calibri", size=10, bold=escolhida)
        if fundo:
            celula.fill = fundo

    aba.cell(row=linha, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    for coluna in (6, 7, 8):
        aba.cell(row=linha, column=coluna).number_format = MOEDA_4
    custo = aba.cell(row=linha, column=9)
    custo.number_format = MOEDA
    custo.font = Font(name="Calibri", size=10, bold=True, color=VERDE if melhor else "FF16294D")

    marca = aba.cell(row=linha, column=10)
    marca.alignment = Alignment(horizontal="center")
    marca.font = Font(name="Calibri", size=11, bold=True, color=AZUL)

    diferenca = aba.cell(row=linha, column=11)
    diferenca.number_format = MOEDA
    diferenca.font = Font(name="Calibri", size=10, color=AMBAR)


# --------------------------------------------------------------------------


def _para_formula(item):
    """Uma linha do banco no formato que `formulas.py` entende. Duplicado de
    `Cotacao.para_calculo` de propósito: aqui o item já veio de um
    `prefetch_related`, e chamar o método do model refaria a consulta."""

    from .formulas import Item, Oferta

    return Item(
        descricao=item.descricao,
        quantidade=item.quantidade,
        margem_minima=item.margem_minima,
        margem_maxima=item.margem_maxima,
        impostos=item.impostos,
        ofertas=[
            Oferta(
                identificador=oferta.pk,
                nome=oferta.nome,
                custo_produto=oferta.custo_produto,
                frete=oferta.frete,
                outros=oferta.outros,
                escolhida=oferta.escolhida,
            )
            for oferta in item.ofertas.all()
        ],
    )


def _configurar_impressao(aba: Worksheet) -> None:
    """A proposta é impressa/virada em PDF com frequência — sem isto sai
    quebrada em cinco páginas de largura."""

    aba.page_setup.orientation = "landscape"
    aba.page_setup.fitToWidth = 1
    aba.page_setup.fitToHeight = 0
    aba.sheet_properties.pageSetUpPr.fitToPage = True
    aba.print_options.horizontalCentered = True
    aba.oddFooter.right.text = "Página &P de &N"
    aba.oddFooter.left.text = "Inside Solutions · proposta gerada pelo Cotador"
